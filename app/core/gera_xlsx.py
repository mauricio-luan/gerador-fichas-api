from app.schemas.Ficha import Ficha
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.drawing.image import Image
from pathlib import Path

# import logging
# logger = logging.getLogger("gerador-fichas-3.0.gera_planilha")


def gera_planilha(ficha: Ficha) -> Workbook:
    """Gera a ficha de implantação com os dados do cliente"""

    wb = Workbook()
    ws = wb.active
    ws.title = "Ficha"

    fonte_negrito = Font(bold=True)
    fonte_branca = Font(color="FFFFFF", bold=True)
    fonte_cinza_claro = Font(color="808080", bold=True)
    alinhamento_central = Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )
    alinhamento_esquerda = Alignment(
        horizontal="left", vertical="center", wrap_text=True
    )
    borda_fina = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    preenchimento_cabecalho = PatternFill(
        start_color="171717", end_color="171717", fill_type="solid"
    )
    preenchimento_secoes = PatternFill(
        start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
    )

    ws.column_dimensions["A"].width = 17
    ws.column_dimensions["B"].width = 50

    ws["A1"].value = "PAYER"
    ws["A1"].font = fonte_branca
    ws["A1"].alignment = alinhamento_central
    ws["A1"].fill = preenchimento_cabecalho
    ws["A1"].border = borda_fina

    DIRETORIO_ATUAL = Path(__file__).parent
    CAMINHO_IMAGEM = DIRETORIO_ATUAL / "payer.png"

    img = Image(CAMINHO_IMAGEM)
    img.width = 120
    img.height = 19
    ws.add_image(img, "A1")

    ws["B1"].value = "FICHA DE IMPLANTAÇÃO"
    ws["B1"].font = fonte_branca
    ws["B1"].alignment = alinhamento_central
    ws["B1"].fill = preenchimento_cabecalho
    ws["B1"].border = borda_fina

    linha_atual = 2
    for chave, valor in ficha.model_dump().items():
        if chave == "account":
            ws.merge_cells(f"A{linha_atual}:B{linha_atual}")
            ws[f"A{linha_atual}"].value = "DADOS PAYER"
            ws[f"A{linha_atual}"].font = fonte_negrito
            ws[f"A{linha_atual}"].alignment = alinhamento_central
            ws[f"A{linha_atual}"].fill = preenchimento_secoes
            ws[f"A{linha_atual}"].border = borda_fina
            linha_atual += 1

        ws[f"A{linha_atual}"].value = chave.upper()
        ws[f"A{linha_atual}"].font = fonte_cinza_claro
        ws[f"A{linha_atual}"].border = borda_fina
        ws[f"A{linha_atual}"].alignment = alinhamento_central

        ws[f"B{linha_atual}"].value = valor
        ws[f"B{linha_atual}"].border = borda_fina
        ws[f"B{linha_atual}"].alignment = alinhamento_esquerda

        if linha_atual in range(13, 18):
            ws[f"B{linha_atual}"].alignment = alinhamento_central
            ws.column_dimensions["B"].width = 48

        linha_atual += 1

    return wb
